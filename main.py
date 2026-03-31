mainimport telebot
from openpyxl import Workbook, load_workbook
from datetime import datetime
import os
import time

TOKEN = "8742759292:AAFrQsqZuasdIJ2HF6PC_yIyG0O9UHly4Q4"
ADMIN_ID = 261124452

bot = telebot.TeleBot(TOKEN)

FILE_NAME = "natijalar.xlsx"

# Excel yaratish
if not os.path.exists(FILE_NAME):
    wb = Workbook()
    ws = wb.active
    ws.append(["Ism familiya", "Ball", "Natija", "Sana"])
    wb.save(FILE_NAME)

# ✅ JAVOBLAR (TEPADA BO‘LISHI SHART)
answers = {
    1: "A", 2: "C", 3: "D", 4: "A", 5: "B",
    6: "C", 7: "C", 8: "C", 9: "A", 10: "A",
    11: "B", 12: "C", 13: "A", 14: "B", 15: "D",
    16: "HAQIQAT QALAMI",
    17: "O'N TO'RT",
    18: "TEMUR MALIK",
    19: "MIRKOMILBOYNING QAZO BO'LGAN NAMOZI",
    20: "1- PARDASIDA"
}

user_data = {}

# START
@bot.message_handler(commands=['start'])
def start(message):
    bot.send_message(message.chat.id, "Ism familiyangizni kiriting:")
    user_data[message.chat.id] = {"step": "name"}

# Excel yuborish
@bot.message_handler(commands=['excel'])
def send_excel(message):
    if message.chat.id == ADMIN_ID:
        with open(FILE_NAME, "rb") as f:
            bot.send_document(message.chat.id, f)
    else:
        bot.send_message(message.chat.id, "Sizda ruxsat yo‘q ❌")

# Asosiy handler
@bot.message_handler(func=lambda m: True)
def main_handler(message):
    chat_id = message.chat.id
    text = message.text.strip()

    if chat_id in user_data and user_data[chat_id]["step"] == "name":
        user_data[chat_id]["name"] = text
        user_data[chat_id]["step"] = "test"
        bot.send_message(chat_id, "Javoblarni yuboring:\n1 A 2 B ... 20")
        return

    if chat_id in user_data and user_data[chat_id]["step"] == "test":

        user_text = text.upper().split()
        user_answers = {}

        i = 0
        while i < len(user_text)-1:
            if user_text[i].isdigit():
                q = int(user_text[i])
                a = user_text[i+1]
                user_answers[q] = a
                i += 2
            else:
                i += 1

        score = 0
        result_text = []

        # ✅ HAR SAVOLNI TEKSHIRISH
        for q, correct in answers.items():
            user_ans = user_answers.get(q, "")

            if user_ans == correct or correct in user_ans:
                score += 1
                result_text.append(f"{q}✅")
            else:
                result_text.append(f"{q}❌")

        natija_string = " ".join(result_text)

        name = user_data[chat_id]["name"]

        # Excelga yozish
        wb = load_workbook(FILE_NAME)
        ws = wb.active
        ws.append([
            name,
            score,
            natija_string,
            datetime.now().strftime("%Y-%m-%d %H:%M")
        ])
        wb.save(FILE_NAME)

        bot.send_message(chat_id, f"{name}\nNatija: {score}/20 ✅")

        del user_data[chat_id]

# DOIMIY ISHLASH
while True:
    try:
        print("Bot ishlayapti...")
        bot.infinity_polling(timeout=60, long_polling_timeout=60)
    except Exception as e:
        print("Xatolik:", e)
        time.sleep(5)
